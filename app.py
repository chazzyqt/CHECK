from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Dict, Union, List, Optional

import cv2
import numpy as np
import json
import asyncio
import concurrent.futures
import logging
import requests
import io
import uvicorn
import aiohttp


from PIL import Image, ImageDraw, ImageFont
from io import BytesIO

import cloudinary
import cloudinary.api

from Final_marker_detection_v3 import MarkerDetector
from parse_answer_full import OCRPipeline
from stringToJSONClass import ExamDataProcessor
from parsedata import UnifiedDataProcessor
from checkscores import TestGrader

import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage


#from openpyxl import Workbook
# Initialize logging
logging.basicConfig(level=logging.INFO)
import base64
# FastAPI app initialization
app = FastAPI()
ocr_pipeline = OCRPipeline()


# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class ImageData(BaseModel):
    original_image: str
    name_section_url: str
    answer_section_url: str
    error: Optional[str] = None

class ImageDataPipeline(BaseModel):
    original_image: str


class ImageRequest(BaseModel):
    uid: str
    data: Dict[str, ImageData]

class RequestModel(BaseModel):
    status: str
    data: Union[Dict[str, ImageData], List[Dict[str, ImageData]]]
    uid: str 

class QuestionAnswerPair(BaseModel):
    question_number: int
    answer: str

class Test(BaseModel):
    test_number: int
    test_type: str
    total_points: int
    correct_points: Optional[float] = None
    question_answer_pairs: List[QuestionAnswerPair]

class StudentInfo(BaseModel):
    university: str
    college: str
    department: str
    exam_type: str
    subject_code: str
    subject_name: str
    name: str
    date: str
    program_code: str
    total_score: Optional[float] = None

class QuestionPair(BaseModel):
    image_name: str
    student_info: StudentInfo
    tests: List[Test]

class RequestModelGrade(BaseModel):
    answer_key: dict
    test_papers: List[dict]

class QuestionAnswerPair(BaseModel):
    question_number: int
    answer: str

class Test(BaseModel):
    test_number: int
    test_type: str
    total_points: int
    correct_points: float
    question_answer_pairs: List[QuestionAnswerPair]

class StudentInfo(BaseModel):
    university: str
    college: str
    department: str
    exam_type: str
    subject_code: str
    subject_name: str
    name: str
    date: str
    program_code: str
    total_score: float

class QuestionPair(BaseModel):
    image_name: str
    student_info: StudentInfo
    tests: List[Test]

class RequestModelUpdatedGrade(BaseModel):
    updated_papers: List[dict]

# Initialize the MarkerDetector
detector = MarkerDetector()

# Maximum file size (e.g., 5MB per file)
MAX_FILE_SIZE = 5 * 1024 * 1024

# Helper function to process a single image (synchronous)
def process_single_image(image_file: UploadFile):
    # Validate the file type
    if image_file.content_type not in ["image/jpeg", "image/png", "image/jpg"]:
        raise HTTPException(status_code=400, detail=f"Unsupported file type: {image_file.content_type}")

    # Read image data and check size
    image_data = image_file.file.read(MAX_FILE_SIZE + 1)
    if len(image_data) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"File {image_file.filename} is too large")

    # Convert image data to OpenCV format
    image_array = np.frombuffer(image_data, np.uint8)
    image = cv2.imdecode(image_array, cv2.IMREAD_COLOR)

    # Check if the image is valid
    if image is None:
        raise HTTPException(status_code=400, detail=f"Invalid image data for {image_file.filename}")

    # Process the image with MarkerDetector
    try:
        result = detector.process_image(image, image_file.filename)
        return json.loads(result)  # Assuming the result is JSON-serializable
    except Exception as e:
        logging.error(f"Error processing image {image_file.filename}: {e}")
        raise HTTPException(status_code=500, detail=f"Error processing image {image_file.filename}")

def convert_to_dataframe_small(updated_papers):
    data = []
    
    for paper in updated_papers:
        question_pair = paper["Question_pair"]
        student_info = question_pair["student_info"]
        
        # Initialize a dictionary for student info
        student_data = {
            "student_name": student_info["name"],
            "college": student_info["college"],
            "department": student_info["department"],
            "exam_type": student_info["exam_type"],
            "subject_code": student_info["subject_code"],
        }
        
        # Loop through each test and record the correct points
        total_score = 0
        for i, test in enumerate(question_pair["tests"], start=1):
            correct_points = test["correct_points"]
            total_score += correct_points  # Accumulate total points
            # Add each test's correct points as a separate column
            student_data[f"Test {i} {test['test_type']}"] = correct_points
        
        # Add total score to the student data
        student_data["Total Score"] = total_score
        
        # Append to the data list
        data.append(student_data)
    
    # Create the dataframe with dynamically generated columns
    df = pd.DataFrame(data)
    return df


# Endpoint to return the Excel file
@app.post("/export_grades_with_images")
async def export_grades_with_images(request: RequestModelUpdatedGrade):
    """Export grades with annotated images in an Excel file."""
    # Convert the updated papers into a DataFrame
    df = convert_to_dataframe_small(request.updated_papers)

    # Save the DataFrame to a BytesIO buffer in Excel format
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write the Grades DataFrame to the first sheet
        df.to_excel(writer, index=False, sheet_name="Grades")
    output.seek(0)

    # Load the workbook to add the Annotated Images sheet
    wb = openpyxl.load_workbook(output)
    annotated_sheet = wb.create_sheet("Annotated Images")

    # Set image cell size constraints
    max_width = 300 * 5
    max_height = 200 * 5

    # Fonts for annotations
    font_path = "JetBrainsMono-Medium.ttf"
    font = ImageFont.truetype(font_path, size=20)
    fontAnnot = ImageFont.truetype("SEGUIEMJ.TTF", size=16)

    # Add images to the Annotated Images sheet
    for idx, paper in enumerate(request.updated_papers):
        original_url = paper["original_url"]
        total_score = paper["Question_pair"]["student_info"]["total_score"]
        tests = paper["Question_pair"]["tests"]

        # Fetch the image
        response = requests.get(original_url)
        img = Image.open(BytesIO(response.content))
        draw = ImageDraw.Draw(img)

        # Annotate with question-answer pairs
        for test in tests:
            for pair in test["question_answer_pairs"]:
                outline_color = "green" if pair["is_correct"] else "blue"
                unicode_symbol = "\u2714" if pair["is_correct"] else "\u2716"
                draw.rectangle(pair["annotation"], outline=outline_color, width=4)
                draw.text(
                    (pair["annotation"][0] - 40, pair["annotation"][1]),
                    unicode_symbol,
                    fill=(0, 128, 0) if pair["is_correct"] else (255, 0, 0),
                    font=fontAnnot,
                )

        # Annotate with full_data annotations
        test_iterator = iter(tests)
        for annotation in paper["full_data"]:
            box = annotation["box"]
            text = annotation["text"]
            draw.rectangle(box, outline=(0, 105, 0), width=2)
            draw.text((box[0] + 10, box[1] - 20), text, fill=(139, 0, 0), font=fontAnnot)

            # Add total score
            if "SCORE:" in text or "SCORE" in text:
                score_position = (box[2] + img.width // 15, box[1] - 30)
                draw.text(score_position, str(int(total_score)), fill="blue", font=font)

            # Add points for specific questions
            if "pts." in text:
                try:
                    test = next(test_iterator)
                    pts_position = (box[2] + img.width // 8, box[1] - 30)
                    draw.text(pts_position, str(int(test["correct_points"])), fill="blue", font=font)
                except StopIteration:
                    pass

        # Resize the image for fitting into the Excel cell
        img.thumbnail((max_width, max_height))
        img_byte_arr = io.BytesIO()
        img.save(img_byte_arr, format="PNG")
        img_byte_arr.seek(0)
        excel_img = ExcelImage(img_byte_arr)

        # Place the image in the Annotated Images sheet
        row_number = idx + 2  # Start from the second row (row 1 is the header)
        annotated_sheet.add_image(excel_img, f"A{row_number}")

        # Add accompanying text data (e.g., URL, Total Score)
        annotated_sheet.cell(row=row_number, column=2, value=original_url)
        annotated_sheet.cell(row=row_number, column=3, value=total_score)

    # Save the updated workbook to the BytesIO buffer
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # Return the Excel file as a response
    return StreamingResponse(
        final_output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=final_grades_with_images.xlsx"},
    )



@app.post("/export_grades_no_images")
async def export_grades_no_images(request: RequestModelUpdatedGrade):
    # Convert the updated papers into a DataFrame
    df = convert_to_dataframe_small(request.updated_papers)



    # Create a DataFrame for annotations
    annotations_df = pd.DataFrame(df)

    # Save the DataFrame to a BytesIO buffer in Excel format
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Grades")
        annotations_df.to_excel(writer, index=False, sheet_name="Annotations")
    output.seek(0)

    # Return the Excel file as a response
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=final_grades_data.xlsx"},
    )


@app.post("/grade_papers")
async def grade_papers(request: RequestModelGrade):
    # Instantiate TestGrader with the provided data
    grader = TestGrader(request.answer_key, request.test_papers)

    # Grade the papers
    grader.grade_papers()

    # Get the updated papers
    updated_papers = grader.get_updated_papers()

    # Return the updated papers
    return {"updated_papers": updated_papers}



async def fetch_image(session, url):
    """Fetch image from URL asynchronously."""
    async with session.get(url) as response:
        return await response.read()

async def fetch_images(image_urls):
    """Fetch all images in parallel."""
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_image(session, url) for url in image_urls]
        return await asyncio.gather(*tasks)



def process_image(image_data, font, fontAnnot):
    """Process a single image and annotate it."""
    original_url = image_data['original_url']
    image_size = image_data["original_size"]
    total_score = image_data["Question_pair"]["student_info"]["total_score"]
    tests = image_data["Question_pair"]["tests"]

    # Open image
    img = Image.open(BytesIO(image_data['content']))
    draw = ImageDraw.Draw(img)

    for test in tests:
        for pair in test["question_answer_pairs"]:
            outline_color = "green" if pair["is_correct"] else "blue"
            unicode_symbol = "\u2714" if  pair["is_correct"] else "\u2716"
            draw.rectangle(pair["annotation"], outline=outline_color, width=4)
            draw.text((pair["annotation"][0] -40, pair["annotation"][1]), unicode_symbol, fill=(0, 128, 0) if pair["is_correct"] else (255, 0, 0), font=fontAnnot)
    # Draw full_data annotations
    test_iterator = iter(tests)
    for annotation in image_data["full_data"]:
        box = annotation['box']
        text = annotation['text']
        draw.rectangle(box, outline=(0, 105, 0), width=2)
        draw.text((box[0] + 10, box[1] - 20), text, fill=(139, 0, 0), font=fontAnnot)

        # Add total score
        if "SCORE:" in text or "SCORE" in text:
            score = str(int(total_score))
            score_position = (box[2] + image_size[0] // 15, box[1] - 30)
            draw.text(score_position, score, fill="blue", font=font)

        # Add test score
        if "pts." in text:
            try:
                test = next(test_iterator)
                pts_score = str(int(test['correct_points']))
                pts_position = (box[2] + image_size[0] // 8, box[1] - 30)
                draw.text(pts_position, pts_score, fill="blue", font=font)
            except StopIteration:
                break

    # Convert image to base64
    buffered = BytesIO()
    img.save(buffered, format="JPEG", optimize=True)
    return base64.b64encode(buffered.getvalue()).decode("utf-8")

@app.post("/get_annotated_image_base_64")
async def get_annotated_image64(request:  RequestModelUpdatedGrade):
    images_base64 = []

    # Preload fonts once
    font = ImageFont.truetype("JetBrainsMono-Medium.ttf", size=40)
    fontAnnot = ImageFont.truetype("SEGUIEMJ.TTF", size=24)

    # Fetch all images asynchronously
    image_urls = [paper['original_url'] for paper in request.updated_papers]
    image_contents = await fetch_images(image_urls)

    # Combine fetched content with original data
    for idx, (image_data, content) in enumerate(zip(request.updated_papers, image_contents)):
        image_data['content'] = content  # Add fetched image content to data
        images_base64.append(process_image(image_data, font, fontAnnot))

    return {"images": images_base64}






# Route to process images (single or multiple)
@app.post("/process_images")
async def process_images(images: list[UploadFile] = File(...), uid: str = Form(...)):
    print(f"UID: {uid}")
    print(f"Images: {images}")

    if not images:
        raise HTTPException(status_code=400, detail="No images provided")

    # Process images concurrently
    loop = asyncio.get_event_loop()
    with concurrent.futures.ThreadPoolExecutor() as executor:
        tasks = [loop.run_in_executor(executor, process_single_image, image) for image in images]
        results = await asyncio.gather(*tasks, return_exceptions=True)  # Run tasks concurrently

    # Check for exceptions in results and add uid
    processed_results = []
    for result in results:
        if isinstance(result, Exception):
            logging.error(f"Processing error: {result}")
            processed_results.append({"status": "error", "error": str(result)})
        else:
            # Add the UID to the result
            
            processed_results.append({"status": "success", "data": result,"uid": uid})

    print(processed_results)
    return JSONResponse(content=processed_results)



@app.post("/delete-images")
async def delete_images(request: RequestModel):
    if request.status != 'success':
        raise HTTPException(status_code=400, detail="Invalid status")
    
    # If the data is wrapped in an array, unwrap it
    if isinstance(request.data, list):
        if len(request.data) > 1:
            raise HTTPException(status_code=400, detail="Invalid format: More than one image set.")
        request.data = request.data[0]  # Unwrap the list to get the first item

    responses = {}
    for image_key, image_data in request.data.items():
        urls = [
            image_data.original_image,
            image_data.name_section_url,
            image_data.answer_section_url
        ]
        
        for url in urls:
            # Extract public_id from URL (assuming format ends with "public_id.jpg")
            public_id = '/'.join(url.split('/')[-2:]).split('.')[0]  # e.g., CHECK/average_lighting.jpg_original
            try:
                response = cloudinary.api.delete_resources([public_id])
                responses[public_id] = response
            except Exception as e:
                responses[public_id] = {"error": str(e)}
    
    return {"deletion_responses": responses}


@app.post("/parse-images")
async def process_images(requests: List[ImageRequest]):
    """
    Endpoint to process images with OCRPipeline.
    """
    try:
        # Clear previous results to ensure no stale data is carried forward
        ocr_pipeline.results = []

        # Convert request objects into a list of dictionaries for processing
        image_data_list = [request.model_dump() for request in requests]
        
        # Process images using the OCR pipeline
        results = ocr_pipeline.process_images(image_data_list)
        data_processor = UnifiedDataProcessor(results)
        all_results = data_processor.process_all_full_entries()
        processor = ExamDataProcessor(all_results)
        processor.process()
        question_pair = processor.get_output_json()
        for i, pair in enumerate(question_pair):
            all_results[i]["Question_pair"] = pair
        #print(all_results)
        return {"status": "success", "results": all_results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
async def health_check():
    return {"status": "healthy"}

# Test route to verify CORS
@app.get("/test")
async def test():
    return {"message": "CORS is working!"}

